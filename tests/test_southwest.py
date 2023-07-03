from chargemaster_parsers.parsers import SouthwestChargeMasterParser, ChargeMasterEntry

import tempfile
from openpyxl import Workbook
import pytest
import io
import os


@pytest.fixture
def parser():
    yield SouthwestChargeMasterParser()


def build_test_workbook(values, scramble=False, include_junk_rows=True):
    wb = Workbook()
    ws = wb.active
    curr_row = 1

    if include_junk_rows:
        ws.cell(
            row=curr_row, column=1, value="Hospital Name: Southwest Healthcare System"
        )
        curr_row += 1
        ws.cell(row=curr_row, column=1, value="Price Effective Date: 4/1/2023")
        curr_row += 1

    if scramble:
        new_values = {}
        for key, value in reversed(values.items()):
            new_values[key] = value
        values = new_values

    # Header row
    for index, key in enumerate(values):
        ws.cell(row=curr_row, column=index + 1, value=key)
    curr_row += 1

    for index, value in enumerate(values.values()):
        ws.cell(row=curr_row, column=index + 1, value=value)
    curr_row += 1

    return wb


TEST_CASE_1 = {
    "Facility": "Southwest Healthcare System",
    "Description": "COMPONENT FEM CR LT 4N",
    "CDM": 38000501,
    "Code Type": "Chargemaster",
    "DRG (If Applicable)": "",
    "CPT/HCPCS (If Applicable)": "C1776",
    "EAPG (If Applicable)": "",
    "APC (If Applicable)": "",
    "Rev Code (If Applicable)": 278,
    "Gross Charge": 1922,
    "Cash Price": 769,
    "Minimum": 398,
    "Maximum": 5419,
    "Aetna HMO/PPO": 848,
    "Aetna Medicare": -1,
    "Blue Cross Anthem": 761,
    "Blue Cross Medi-Cal": -1,
    "Blue Cross Senior": -1,
    "Blue Shield California Promise": -1,
    "Blue Shield Promise": -1,
    "Blue Shield Promise Rady": -1,
    "Blue Shield Select": -1,
    "Blue Shield Senior": -1,
    "Brand New Day": -1,
    "Cal Optima Medicaid": -1,
    "Cigna HMO/PPO": 550,
    "Epic Health": 2956,
    "Epic Health Plan Medicare": -1,
    "Exclusive Care": 3941,
    "First Health": 884,
    "HealthNet": 446,
    "HealthNet Medi-Cal": -1,
    "HealthNet Medicare": 4050,
    "Heritage Commercial": 475,
    "Heritage Medi-Cal": -1,
    "Heritage Medicare": 398,
    "Humana Medicare": -1,
    "Inland Empire Health Plan": -1,
    "Inland Empire Health Plan Medicare": -1,
    "Kaiser": 646,
    "Kaiser Medi-Cal": -1,
    "Kaiser Medicare": -1,
    "Molina": -1,
    "Molina Medi-Cal": -1,
    "Molina Medicare": -1,
    "Multiplan": 1730,
    "Palomar Health": 942,
    "Scan Medicare": -1,
    "Sharp Health Plan": 5419,
    "United Healthcare HMO": -1,
    "United Healthcare PPO": -1,
    "United Healthcare Community Plan": -1,
    "United Healthcare Medicare": -1,
}

EXPECTED_RESULTS_1 = [
    ChargeMasterEntry(
        expected_reimbursement=848,
        gross_charge=1922.0,
        hcpcs_code="C1776",
        max_reimbursement=5419.0,
        min_reimbursement=398.0,
        payer="Aetna HMO/PPO",
        procedure_description="COMPONENT FEM CR LT 4N",
        procedure_identifier=38000501,
    ),
    ChargeMasterEntry(
        expected_reimbursement=761,
        gross_charge=1922.0,
        hcpcs_code="C1776",
        max_reimbursement=5419.0,
        min_reimbursement=398.0,
        payer="Blue Cross Anthem",
        procedure_description="COMPONENT FEM CR LT 4N",
        procedure_identifier=38000501,
    ),
    ChargeMasterEntry(
        expected_reimbursement=550,
        gross_charge=1922.0,
        hcpcs_code="C1776",
        max_reimbursement=5419.0,
        min_reimbursement=398.0,
        payer="Cigna HMO/PPO",
        procedure_description="COMPONENT FEM CR LT 4N",
        procedure_identifier=38000501,
    ),
    ChargeMasterEntry(
        expected_reimbursement=2956,
        gross_charge=1922.0,
        hcpcs_code="C1776",
        max_reimbursement=5419.0,
        min_reimbursement=398.0,
        payer="Epic Health",
        procedure_description="COMPONENT FEM CR LT 4N",
        procedure_identifier=38000501,
    ),
    ChargeMasterEntry(
        expected_reimbursement=3941,
        gross_charge=1922.0,
        hcpcs_code="C1776",
        max_reimbursement=5419.0,
        min_reimbursement=398.0,
        payer="Exclusive Care",
        procedure_description="COMPONENT FEM CR LT 4N",
        procedure_identifier=38000501,
    ),
    ChargeMasterEntry(
        expected_reimbursement=884,
        gross_charge=1922.0,
        hcpcs_code="C1776",
        max_reimbursement=5419.0,
        min_reimbursement=398.0,
        payer="First Health",
        procedure_description="COMPONENT FEM CR LT 4N",
        procedure_identifier=38000501,
    ),
    ChargeMasterEntry(
        expected_reimbursement=446,
        gross_charge=1922.0,
        hcpcs_code="C1776",
        max_reimbursement=5419.0,
        min_reimbursement=398.0,
        payer="HealthNet",
        procedure_description="COMPONENT FEM CR LT 4N",
        procedure_identifier=38000501,
    ),
    ChargeMasterEntry(
        expected_reimbursement=4050,
        gross_charge=1922.0,
        hcpcs_code="C1776",
        max_reimbursement=5419.0,
        min_reimbursement=398.0,
        payer="HealthNet Medicare",
        procedure_description="COMPONENT FEM CR LT 4N",
        procedure_identifier=38000501,
    ),
    ChargeMasterEntry(
        expected_reimbursement=475,
        gross_charge=1922.0,
        hcpcs_code="C1776",
        max_reimbursement=5419.0,
        min_reimbursement=398.0,
        payer="Heritage Commercial",
        procedure_description="COMPONENT FEM CR LT 4N",
        procedure_identifier=38000501,
    ),
    ChargeMasterEntry(
        expected_reimbursement=398,
        gross_charge=1922.0,
        hcpcs_code="C1776",
        max_reimbursement=5419.0,
        min_reimbursement=398.0,
        payer="Heritage Medicare",
        procedure_description="COMPONENT FEM CR LT 4N",
        procedure_identifier=38000501,
    ),
    ChargeMasterEntry(
        expected_reimbursement=646,
        gross_charge=1922.0,
        hcpcs_code="C1776",
        max_reimbursement=5419.0,
        min_reimbursement=398.0,
        payer="Kaiser",
        procedure_description="COMPONENT FEM CR LT 4N",
        procedure_identifier=38000501,
    ),
    ChargeMasterEntry(
        expected_reimbursement=1730,
        gross_charge=1922.0,
        hcpcs_code="C1776",
        max_reimbursement=5419.0,
        min_reimbursement=398.0,
        payer="Multiplan",
        procedure_description="COMPONENT FEM CR LT 4N",
        procedure_identifier=38000501,
    ),
    ChargeMasterEntry(
        expected_reimbursement=942,
        gross_charge=1922.0,
        hcpcs_code="C1776",
        max_reimbursement=5419.0,
        min_reimbursement=398.0,
        payer="Palomar Health",
        procedure_description="COMPONENT FEM CR LT 4N",
        procedure_identifier=38000501,
    ),
    ChargeMasterEntry(
        expected_reimbursement=5419,
        gross_charge=1922.0,
        hcpcs_code="C1776",
        max_reimbursement=5419.0,
        min_reimbursement=398.0,
        payer="Sharp Health Plan",
        procedure_description="COMPONENT FEM CR LT 4N",
        procedure_identifier=38000501,
    ),
    ChargeMasterEntry(
        gross_charge=769.0,
        hcpcs_code="C1776",
        payer="Cash",
        procedure_description="COMPONENT FEM CR LT 4N",
        procedure_identifier=38000501,
    ),
]


def test_simple_row(parser):
    wb = build_test_workbook(TEST_CASE_1)

    expected_result = EXPECTED_RESULTS_1

    with tempfile.TemporaryDirectory() as tmp_dir:
        filename = os.path.join(tmp_dir, "southwest.xlsx")
        wb.save(filename)
        actual_result = list(
            parser.parse_artifacts(
                {SouthwestChargeMasterParser.ARTIFACT_URL: open(filename, "rb")}
            )
        )

    assert sorted(expected_result) == sorted(actual_result)

def test_simple_messed_up(parser):
    wb = build_test_workbook(TEST_CASE_1, scramble=True, include_junk_rows=False)

    expected_result = EXPECTED_RESULTS_1

    with tempfile.TemporaryDirectory() as tmp_dir:
        filename = os.path.join(tmp_dir, "southwest.xlsx")
        wb.save(filename)
        actual_result = list(
            parser.parse_artifacts(
                {SouthwestChargeMasterParser.ARTIFACT_URL: open(filename, "rb")}
            )
        )

    assert sorted(expected_result) == sorted(actual_result)

def test_extra_apc(parser):
    test_case = {"Facility": "Southwest Healthcare System",
    "Description": "PATHOGEN REDUCED PLATELETS",
    "CDM": 50302033,
    "Code Type": "Chargemaster",
    "DRG (If Applicable)": "",
    "CPT/HCPCS (If Applicable)": "P9073",
    "EAPG (If Applicable)": "",
    "APC (If Applicable)": "9536",
    "Rev Code (If Applicable)": 390,
    "Gross Charge": 793,
    "Cash Price": 317,
    "Minimum": 266,
    "Maximum": 974,
    "Aetna HMO/PPO": 288}

    expected_result = [
        ChargeMasterEntry(
            expected_reimbursement=288,
            gross_charge=793.0,
            hcpcs_code="P9073",
            max_reimbursement=974.0,
            min_reimbursement=266.0,
            payer="Aetna HMO/PPO",
            procedure_description="PATHOGEN REDUCED PLATELETS",
            procedure_identifier=50302033,
            extra_data = {"APC (If Applicable)" : "9536"}
        ),
        ChargeMasterEntry(
            gross_charge=317,
            hcpcs_code="P9073",
            payer="Cash",
            procedure_description="PATHOGEN REDUCED PLATELETS",
            procedure_identifier=50302033,
            extra_data = {"APC (If Applicable)" : "9536"}
        ),
    ]

    wb = build_test_workbook(test_case, scramble=True, include_junk_rows=False)

    expected_result = expected_result

    with tempfile.TemporaryDirectory() as tmp_dir:
        filename = os.path.join(tmp_dir, "southwest.xlsx")
        wb.save(filename)
        actual_result = list(
            parser.parse_artifacts(
                {SouthwestChargeMasterParser.ARTIFACT_URL: open(filename, "rb")}
            )
        )

    assert sorted(expected_result) == sorted(actual_result)

def test_extra_apc(parser):
    test_case = {"Facility": "Southwest Healthcare System",
    "Description": "PATHOGEN REDUCED PLATELETS",
    "CDM": 50302033,
    "Code Type": "Chargemaster",
    "DRG (If Applicable)": "",
    "CPT/HCPCS (If Applicable)": "P9073",
    "EAPG (If Applicable)": "",
    "APC (If Applicable)": "9536",
    "Rev Code (If Applicable)": 390,
    "Gross Charge": 793,
    "Cash Price": 317,
    "Minimum": 266,
    "Maximum": 974,
    "Aetna HMO/PPO": 288}

    expected_result = [
        ChargeMasterEntry(
            expected_reimbursement=288,
            gross_charge=793.0,
            hcpcs_code="P9073",
            max_reimbursement=974.0,
            min_reimbursement=266.0,
            payer="Aetna HMO/PPO",
            procedure_description="PATHOGEN REDUCED PLATELETS",
            procedure_identifier=50302033,
            extra_data = {"APC (If Applicable)" : "9536"}
        ),
        ChargeMasterEntry(
            gross_charge=317,
            hcpcs_code="P9073",
            payer="Cash",
            procedure_description="PATHOGEN REDUCED PLATELETS",
            procedure_identifier=50302033,
            extra_data = {"APC (If Applicable)" : "9536"}
        ),
    ]

    wb = build_test_workbook(test_case, scramble=True, include_junk_rows=False)

    expected_result = expected_result

    with tempfile.TemporaryDirectory() as tmp_dir:
        filename = os.path.join(tmp_dir, "southwest.xlsx")
        wb.save(filename)
        actual_result = list(
            parser.parse_artifacts(
                {SouthwestChargeMasterParser.ARTIFACT_URL: open(filename, "rb")}
            )
        )

    assert sorted(expected_result) == sorted(actual_result)

def test_other_cpt(parser):
    test_case = {"Facility": "Southwest Healthcare System",
    "Description": "Spirom fev/fvc>/=70%/w/ocopd",
    "CDM": "",
    "Code Type": "Other CPT/HCPCS",
    "DRG (If Applicable)": "",
    "CPT/HCPCS (If Applicable)": "3027F",
    "EAPG (If Applicable)": "",
    "APC (If Applicable)": "",
    "Rev Code (If Applicable)": "",
    "Gross Charge": -1,
    "Cash Price": -1,
    "Minimum": 10526,
    "Maximum": 10526,
    "Blue Shield Select": 10526
}

    expected_result = [
        ChargeMasterEntry(
            expected_reimbursement=10526.0,
            procedure_identifier="CPT_3027F",
            cpt_code="3027F",
            max_reimbursement=10526.0,
            min_reimbursement=10526.0,
            payer="Blue Shield Select",
            procedure_description="Spirom fev/fvc>/=70%/w/ocopd",
        ),
    ]

    wb = build_test_workbook(test_case, scramble=True, include_junk_rows=False)

    expected_result = expected_result

    with tempfile.TemporaryDirectory() as tmp_dir:
        filename = os.path.join(tmp_dir, "southwest.xlsx")
        wb.save(filename)
        actual_result = list(
            parser.parse_artifacts(
                {SouthwestChargeMasterParser.ARTIFACT_URL: open(filename, "rb")}
            )
        )

    assert sorted(expected_result) == sorted(actual_result)

def test_ms_drg(parser):
    test_case = {"Facility": "Southwest Healthcare System",
    "Description": "HEART TRANSPLANT OR IMPLANT OF HEART ASSIST SYSTEM WITH MCC",
    "CDM": "",
    "Code Type": "MS-DRG",
    "DRG (If Applicable)": 1,
    "CPT/HCPCS (If Applicable)": "",
    "EAPG (If Applicable)": "",
    "APC (If Applicable)": "",
    "Rev Code (If Applicable)": "",
    "Gross Charge": -1,
    "Cash Price": -1,
    "Minimum": 239580,
    "Maximum": 448170,
    "Aetna Medicare": 254017
}    


    expected_result = [
        ChargeMasterEntry(
            expected_reimbursement=254017.0,
            procedure_identifier="MS_DRG_001",
            ms_drg_code="001",
            max_reimbursement=448170.0,
            min_reimbursement=239580.0,
            payer="Aetna Medicare",
            procedure_description="HEART TRANSPLANT OR IMPLANT OF HEART ASSIST SYSTEM WITH MCC",
        ),
    ]

    wb = build_test_workbook(test_case, scramble=True, include_junk_rows=False)

    expected_result = expected_result

    with tempfile.TemporaryDirectory() as tmp_dir:
        filename = os.path.join(tmp_dir, "southwest.xlsx")
        wb.save(filename)
        actual_result = list(
            parser.parse_artifacts(
                {SouthwestChargeMasterParser.ARTIFACT_URL: open(filename, "rb")}
            )
        )

    assert sorted(expected_result) == sorted(actual_result)


def test_institution_name(parser):
    assert SouthwestChargeMasterParser.institution_name == "Southwest"
    assert parser.institution_name == "Southwest"


def test_artifact_urls(parser):
    assert (
        SouthwestChargeMasterParser.artifact_urls
        == SouthwestChargeMasterParser.ARTIFACT_URLS
    )
    assert parser.artifact_urls == SouthwestChargeMasterParser.ARTIFACT_URLS
